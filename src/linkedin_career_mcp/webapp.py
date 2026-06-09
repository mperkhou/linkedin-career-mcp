from __future__ import annotations

import argparse
import sqlite3
import subprocess
import sys
import threading
from dataclasses import dataclass
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_OUTPUT_DIR = Path("output")
TRACKING_WORKBOOK = Path("tracking/read_applications/linkedin_applications.xlsx")
DEFAULT_DATABASE = Path("tracking/applications.sqlite3")

TRACKING_COLUMNS = (
    "job_id",
    "company",
    "job_title",
    "linkedin_url",
    "customized_resume",
    "cover_letter",
    "applied_to",
    "date_applied",
)
APPLICATION_STATUSES = {"No", "Yes", "N/A"}
REQUIRED_TRACKING_COLUMNS = (
    "job_id",
    "company",
    "job_title",
    "linkedin_url",
    "customized_resume",
    "applied_to",
    "date_applied",
)
APPLICATION_EXTRA_COLUMNS = {
    "job_description": "TEXT",
    "prompt_job_description": "TEXT",
    "cover_letter_filename": "TEXT NOT NULL DEFAULT ''",
    "cover_letter_content": "BLOB",
    "cover_letter_mime_type": "TEXT NOT NULL DEFAULT 'application/pdf'",
    "source_cover_letter_path": "TEXT NOT NULL DEFAULT ''",
    "date_matched": "TEXT",
    "date_posted": "TEXT",
}


@dataclass(frozen=True)
class ImportResult:
    rows_seen: int
    rows_imported: int
    missing_resumes: int
    missing_cover_letters: int = 0


@dataclass(frozen=True)
class ApplicationJobRecord:
    job_id: str
    company: str
    job_title: str
    linkedin_url: str
    job_description: str | None
    prompt_job_description: str | None
    date_matched: str | None
    date_posted: str | None


def connect_database(database_path: Path) -> sqlite3.Connection:
    database_path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(database_path)
    connection.row_factory = sqlite3.Row
    init_database(connection)
    return connection


def init_database(connection: sqlite3.Connection) -> None:
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS applications (
            job_id TEXT PRIMARY KEY,
            company TEXT NOT NULL,
            job_title TEXT NOT NULL,
            linkedin_url TEXT NOT NULL,
            job_description TEXT,
            prompt_job_description TEXT,
            resume_filename TEXT NOT NULL,
            resume_content BLOB,
            resume_mime_type TEXT NOT NULL DEFAULT 'application/pdf',
            source_resume_path TEXT NOT NULL,
            cover_letter_filename TEXT NOT NULL DEFAULT '',
            cover_letter_content BLOB,
            cover_letter_mime_type TEXT NOT NULL DEFAULT 'application/pdf',
            source_cover_letter_path TEXT NOT NULL DEFAULT '',
            date_matched TEXT,
            date_posted TEXT,
            applied_to TEXT NOT NULL DEFAULT 'No',
            date_applied TEXT,
            notes TEXT NOT NULL DEFAULT '',
            imported_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    _ensure_application_columns(connection)
    connection.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS applications_unique_linkedin_job_id
        ON applications(job_id)
        """
    )
    connection.commit()


def _ensure_application_columns(connection: sqlite3.Connection) -> None:
    rows = connection.execute("PRAGMA table_info(applications)").fetchall()
    existing_columns = {str(row["name"]) for row in rows}
    for column, column_type in APPLICATION_EXTRA_COLUMNS.items():
        if column not in existing_columns:
            connection.execute(f"ALTER TABLE applications ADD COLUMN {column} {column_type}")
    connection.execute(
        """
        UPDATE applications
        SET date_matched = imported_at
        WHERE (date_matched IS NULL OR TRIM(date_matched) = '')
          AND imported_at IS NOT NULL
        """
    )


def fetch_existing_resume_job_ids(database_path: Path) -> set[str]:
    with connect_database(database_path) as connection:
        rows = connection.execute(
            """
            SELECT job_id
            FROM applications
            WHERE resume_content IS NOT NULL
            """
        ).fetchall()
    return {str(row["job_id"]) for row in rows}


def fetch_existing_cover_letter_job_ids(database_path: Path) -> set[str]:
    with connect_database(database_path) as connection:
        rows = connection.execute(
            """
            SELECT job_id
            FROM applications
            WHERE cover_letter_content IS NOT NULL
            """
        ).fetchall()
    return {str(row["job_id"]) for row in rows}


def fetch_application_job_ids(database_path: Path) -> set[str]:
    with connect_database(database_path) as connection:
        rows = connection.execute("SELECT job_id FROM applications").fetchall()
    return {str(row["job_id"]) for row in rows}


def fetch_application_job_records(
    database_path: Path,
    *,
    job_ids: list[str] | None = None,
) -> list[ApplicationJobRecord]:
    with connect_database(database_path) as connection:
        if job_ids is None:
            rows = connection.execute(
                """
                SELECT job_id, company, job_title, linkedin_url, job_description,
                       prompt_job_description, date_matched, date_posted
                FROM applications
                ORDER BY company COLLATE NOCASE ASC, job_title COLLATE NOCASE ASC
                """
            ).fetchall()
        elif not job_ids:
            rows = []
        else:
            placeholders = ", ".join("?" for _ in job_ids)
            rows = connection.execute(
                f"""
                SELECT job_id, company, job_title, linkedin_url, job_description,
                       prompt_job_description, date_matched, date_posted
                FROM applications
                WHERE job_id IN ({placeholders})
                """,
                job_ids,
            ).fetchall()
            row_by_job_id = {str(row["job_id"]): row for row in rows}
            rows = [row_by_job_id[job_id] for job_id in job_ids if job_id in row_by_job_id]
    return [
        ApplicationJobRecord(
            job_id=str(row["job_id"]),
            company=str(row["company"] or ""),
            job_title=str(row["job_title"] or ""),
            linkedin_url=str(row["linkedin_url"] or ""),
            job_description=row["job_description"],
            prompt_job_description=row["prompt_job_description"],
            date_matched=row["date_matched"],
            date_posted=row["date_posted"],
        )
        for row in rows
    ]


def upsert_application_artifact(
    *,
    database_path: Path,
    job_id: str,
    company: str,
    job_title: str,
    linkedin_url: str,
    resume_path: Path | None,
    cover_letter_path: Path | None = None,
    job_description: str | None = None,
    prompt_job_description: str | None = None,
    applied_to: str = "No",
    date_applied: str | None = None,
    date_matched: str | None = None,
    date_posted: str | None = None,
) -> None:
    now = datetime.now(UTC).isoformat(timespec="seconds")
    applied_to = _normalize_applied_to(applied_to)
    date_matched = _date_value(date_matched) or now
    date_posted = _date_value(date_posted)
    resume_content = (
        resume_path.read_bytes()
        if resume_path is not None and resume_path.is_file()
        else None
    )
    cover_letter_content = (
        cover_letter_path.read_bytes()
        if cover_letter_path is not None and cover_letter_path.is_file()
        else None
    )
    with connect_database(database_path) as connection:
        connection.execute(
            """
            INSERT INTO applications (
                job_id, company, job_title, linkedin_url, job_description,
                prompt_job_description, resume_filename, resume_content, resume_mime_type,
                source_resume_path, cover_letter_filename, cover_letter_content,
                cover_letter_mime_type, source_cover_letter_path, date_matched, date_posted,
                applied_to, date_applied, imported_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(job_id) DO UPDATE SET
                company = excluded.company,
                job_title = excluded.job_title,
                linkedin_url = excluded.linkedin_url,
                job_description = COALESCE(
                    excluded.job_description,
                    applications.job_description
                ),
                prompt_job_description = COALESCE(
                    excluded.prompt_job_description,
                    applications.prompt_job_description
                ),
                resume_filename = CASE
                    WHEN excluded.resume_filename != '' THEN excluded.resume_filename
                    ELSE applications.resume_filename
                END,
                resume_content = COALESCE(excluded.resume_content, applications.resume_content),
                resume_mime_type = CASE
                    WHEN excluded.resume_content IS NOT NULL THEN excluded.resume_mime_type
                    ELSE applications.resume_mime_type
                END,
                source_resume_path = CASE
                    WHEN excluded.source_resume_path != '' THEN excluded.source_resume_path
                    ELSE applications.source_resume_path
                END,
                cover_letter_filename = CASE
                    WHEN excluded.cover_letter_filename != '' THEN excluded.cover_letter_filename
                    ELSE applications.cover_letter_filename
                END,
                cover_letter_content = COALESCE(
                    excluded.cover_letter_content,
                    applications.cover_letter_content
                ),
                cover_letter_mime_type = CASE
                    WHEN excluded.cover_letter_content IS NOT NULL
                    THEN excluded.cover_letter_mime_type
                    ELSE applications.cover_letter_mime_type
                END,
                source_cover_letter_path = CASE
                    WHEN excluded.source_cover_letter_path != ''
                    THEN excluded.source_cover_letter_path
                    ELSE applications.source_cover_letter_path
                END,
                date_matched = COALESCE(
                    NULLIF(applications.date_matched, ''),
                    excluded.date_matched
                ),
                date_posted = COALESCE(
                    NULLIF(excluded.date_posted, ''),
                    applications.date_posted
                ),
                applied_to = CASE
                    WHEN applications.applied_to != 'No' THEN applications.applied_to
                    ELSE excluded.applied_to
                END,
                date_applied = COALESCE(applications.date_applied, excluded.date_applied),
                updated_at = excluded.updated_at
            """,
            (
                str(job_id),
                company,
                job_title,
                linkedin_url,
                job_description,
                prompt_job_description,
                resume_path.name if resume_path is not None else "",
                resume_content,
                "application/pdf",
                str(resume_path) if resume_path is not None else "",
                cover_letter_path.name if cover_letter_path is not None else "",
                cover_letter_content,
                "application/pdf",
                str(cover_letter_path) if cover_letter_path is not None else "",
                date_matched,
                date_posted,
                applied_to,
                date_applied,
                now,
                now,
            ),
        )
        connection.commit()


def delete_applications(
    *,
    database_path: Path,
    job_ids: list[str],
    tracking_path: Path | None = None,
) -> int:
    normalized_job_ids = sorted({job_id.strip() for job_id in job_ids if job_id.strip()})
    if not normalized_job_ids:
        return 0

    placeholders = ", ".join("?" for _ in normalized_job_ids)
    with connect_database(database_path) as connection:
        cursor = connection.execute(
            f"DELETE FROM applications WHERE job_id IN ({placeholders})",
            normalized_job_ids,
        )
        deleted_count = cursor.rowcount
        connection.commit()

    if tracking_path is not None:
        _delete_tracking_rows(tracking_path=tracking_path, job_ids=set(normalized_job_ids))
    return deleted_count


def import_output_artifacts(*, output_dir: Path, database_path: Path) -> ImportResult:
    tracking_path = output_dir / TRACKING_WORKBOOK
    if not tracking_path.exists():
        raise FileNotFoundError(f"Tracking workbook was not found: {tracking_path}")

    connection = connect_database(database_path)
    connection.close()
    workbook = load_workbook(tracking_path, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    column_indexes = {header: index for index, header in enumerate(headers)}
    missing_columns = [
        column for column in REQUIRED_TRACKING_COLUMNS if column not in column_indexes
    ]
    if missing_columns:
        raise ValueError(f"Tracking workbook is missing columns: {', '.join(missing_columns)}")

    rows_seen = 0
    rows_imported = 0
    missing_resumes = 0
    missing_cover_letters = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = {
            column: row[column_indexes[column]]
            for column in REQUIRED_TRACKING_COLUMNS
        }
        job_id = str(values["job_id"] or "").strip()
        if not job_id:
            continue
        rows_seen += 1

        resume_path_text = str(values["customized_resume"] or "").strip()
        resume_path = (
            _resolve_output_path(output_dir=output_dir, path_text=resume_path_text)
            if resume_path_text
            else None
        )
        if resume_path_text and (resume_path is None or not resume_path.exists()):
            missing_resumes += 1

        cover_letter_path_text = ""
        if "cover_letter" in column_indexes:
            cover_letter_value = row[column_indexes["cover_letter"]]
            cover_letter_path_text = str(cover_letter_value or "").strip()
        cover_letter_path = (
            _resolve_output_path(output_dir=output_dir, path_text=cover_letter_path_text)
            if cover_letter_path_text
            else None
        )
        if cover_letter_path_text and (
            cover_letter_path is None or not cover_letter_path.exists()
        ):
            missing_cover_letters += 1

        upsert_application_artifact(
            database_path=database_path,
            job_id=job_id,
            company=str(values["company"] or ""),
            job_title=str(values["job_title"] or ""),
            linkedin_url=str(values["linkedin_url"] or ""),
            resume_path=resume_path,
            cover_letter_path=cover_letter_path,
            applied_to=str(values["applied_to"] or "No"),
            date_applied=_date_value(values["date_applied"]),
            date_matched=_optional_row_value(
                row=row,
                column_indexes=column_indexes,
                column="date_matched",
            ),
            date_posted=_optional_row_value(
                row=row,
                column_indexes=column_indexes,
                column="date_posted",
            ),
        )
        rows_imported += 1

    return ImportResult(
        rows_seen=rows_seen,
        rows_imported=rows_imported,
        missing_resumes=missing_resumes,
        missing_cover_letters=missing_cover_letters,
    )


def create_app(*, database_path: Path, output_dir: Path):
    from flask import (
        Flask,
        Response,
        abort,
        flash,
        redirect,
        render_template_string,
        request,
        send_file,
        url_for,
    )

    app = Flask(__name__)
    app.config["SECRET_KEY"] = "linkedin-career-local-only"
    app.config["DATABASE_PATH"] = database_path
    app.config["OUTPUT_DIR"] = output_dir
    app.jinja_env.filters["display_date"] = _display_date

    @app.get("/")
    def index():
        rows = _fetch_applications(database_path)
        stats = {
            "total": len(rows),
            "applied": sum(1 for row in rows if row["applied_to"] == "Yes"),
            "pending": sum(1 for row in rows if row["applied_to"] == "No"),
            "not_applicable": sum(1 for row in rows if row["applied_to"] == "N/A"),
        }
        return render_template_string(INDEX_TEMPLATE, rows=rows, stats=stats)

    @app.post("/applications/<job_id>")
    def update_application(job_id: str):
        applied_to = _normalize_applied_to(request.form.get("applied_to"))
        date_applied = request.form.get("date_applied") or None
        notes = request.form.get("notes", "")
        now = datetime.now(UTC).isoformat(timespec="seconds")
        with connect_database(database_path) as connection:
            connection.execute(
                """
                UPDATE applications
                SET applied_to = ?, date_applied = ?, notes = ?, updated_at = ?
                WHERE job_id = ?
                """,
                (applied_to, date_applied, notes, now, job_id),
            )
            connection.commit()
        if applied_to == "Yes":
            deleted_count = cleanup_downloaded_application_pdfs()
            flash(f"Application updated. Cleared {deleted_count} downloaded PDF file(s).")
        else:
            flash("Application updated.")
        return redirect(url_for("index"))

    @app.post("/sync")
    def sync_from_output():
        result = import_output_artifacts(output_dir=output_dir, database_path=database_path)
        flash(
            f"Imported {result.rows_imported} workbook rows "
            f"({result.missing_resumes} missing resume files, "
            f"{result.missing_cover_letters} missing cover letter files)."
        )
        return redirect(url_for("index"))

    @app.post("/applications/delete")
    def bulk_delete_applications():
        job_ids = request.form.getlist("job_id")
        deleted_count = delete_applications(
            database_path=database_path,
            job_ids=job_ids,
            tracking_path=output_dir / TRACKING_WORKBOOK,
        )
        flash(f"Deleted {deleted_count} application rows.")
        return redirect(url_for("index"))

    @app.get("/linkedin/<job_id>")
    def open_linkedin_job(job_id: str):
        row = _fetch_application(database_path, job_id)
        if row is None or not row["linkedin_url"]:
            abort(404)
        _open_url_in_chromium(str(row["linkedin_url"]))
        flash("Opened LinkedIn job in Chromium.")
        return redirect(url_for("index"))

    @app.get("/resumes/<job_id>")
    def resume(job_id: str):
        row = _fetch_application(database_path, job_id)
        if row is None or row["resume_content"] is None:
            return Response("Resume was not found in the database.", status=404)
        return send_file(
            BytesIO(row["resume_content"]),
            mimetype=row["resume_mime_type"],
            download_name=row["resume_filename"],
            as_attachment=False,
        )

    @app.get("/resumes/<job_id>/download")
    def resume_download(job_id: str):
        row = _fetch_application(database_path, job_id)
        if row is None or row["resume_content"] is None:
            return Response("Resume was not found in the database.", status=404)
        return send_file(
            BytesIO(row["resume_content"]),
            mimetype=row["resume_mime_type"],
            download_name=row["resume_filename"],
            as_attachment=True,
        )

    @app.get("/cover-letters/<job_id>")
    def cover_letter(job_id: str):
        row = _fetch_application(database_path, job_id)
        if row is None or row["cover_letter_content"] is None:
            return Response("Cover letter was not found in the database.", status=404)
        return send_file(
            BytesIO(row["cover_letter_content"]),
            mimetype=row["cover_letter_mime_type"],
            download_name=row["cover_letter_filename"],
            as_attachment=False,
        )

    @app.get("/cover-letters/<job_id>/download")
    def cover_letter_download(job_id: str):
        row = _fetch_application(database_path, job_id)
        if row is None or row["cover_letter_content"] is None:
            return Response("Cover letter was not found in the database.", status=404)
        return send_file(
            BytesIO(row["cover_letter_content"]),
            mimetype=row["cover_letter_mime_type"],
            download_name=row["cover_letter_filename"],
            as_attachment=True,
        )

    @app.get("/descriptions/<job_id>")
    def compare_descriptions(job_id: str):
        row = _fetch_application(database_path, job_id)
        if row is None:
            abort(404)
        return render_template_string(DESCRIPTION_COMPARE_TEMPLATE, row=row)

    @app.get("/output/<path:relative_path>")
    def output_file(relative_path: str):
        target = (output_dir / relative_path).resolve()
        output_root = output_dir.resolve()
        if target != output_root and output_root not in target.parents:
            abort(404)
        if not target.is_file():
            abort(404)
        return send_file(target, as_attachment=False)

    return app


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Launch the local LinkedIn application tracker.")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--database", default="")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8765)
    parser.add_argument("--debug", action="store_true")
    parser.add_argument("--no-import", action="store_true")
    parser.add_argument("--open-browser", action="store_true")
    args = parser.parse_args(argv)

    output_dir = Path(args.output_dir)
    database_path = Path(args.database) if args.database else output_dir / DEFAULT_DATABASE
    if not args.no_import:
        import_output_artifacts(output_dir=output_dir, database_path=database_path)
    app = create_app(database_path=database_path, output_dir=output_dir)
    if args.open_browser:
        _schedule_browser_open(host=args.host, port=args.port)
    app.run(host=args.host, port=args.port, debug=args.debug)


def _fetch_applications(database_path: Path) -> list[sqlite3.Row]:
    with connect_database(database_path) as connection:
        return list(
            connection.execute(
                """
                SELECT *
                FROM applications
                ORDER BY
                    CASE applied_to
                        WHEN 'No' THEN 0
                        WHEN 'N/A' THEN 1
                        WHEN 'Yes' THEN 2
                        ELSE 3
                    END,
                    company COLLATE NOCASE ASC,
                    job_title COLLATE NOCASE ASC
                """
            )
        )


def _fetch_application(database_path: Path, job_id: str) -> sqlite3.Row | None:
    with connect_database(database_path) as connection:
        return connection.execute(
            "SELECT * FROM applications WHERE job_id = ?",
            (job_id,),
        ).fetchone()


def _resolve_output_path(*, output_dir: Path, path_text: str) -> Path:
    artifact_path = Path(path_text)
    if artifact_path.is_absolute():
        return artifact_path
    if artifact_path.parts and artifact_path.parts[0] == output_dir.name:
        return output_dir.parent / artifact_path
    return output_dir / artifact_path


def _optional_row_value(
    *,
    row: tuple[Any, ...],
    column_indexes: dict[str, int],
    column: str,
) -> Any:
    index = column_indexes.get(column)
    if index is None:
        return None
    return row[index]


def _date_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    return text or None


def _normalize_applied_to(value: Any) -> str:
    text = str(value or "").strip()
    return text if text in APPLICATION_STATUSES else "No"


def cleanup_downloaded_application_pdfs(download_dir: Path | None = None) -> int:
    target_dir = download_dir or Path.home() / "Downloads"
    if not target_dir.is_dir():
        return 0

    deleted_count = 0
    for path in target_dir.glob("mp_*.pdf"):
        if not path.is_file():
            continue
        try:
            path.unlink()
        except OSError:
            continue
        deleted_count += 1
    return deleted_count


def _display_date(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if "T" in text:
        return text.split("T", 1)[0]
    return text


def _delete_tracking_rows(*, tracking_path: Path, job_ids: set[str]) -> None:
    if not tracking_path.exists():
        return

    workbook = load_workbook(tracking_path)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    if "job_id" not in headers:
        return

    job_id_column = headers.index("job_id") + 1
    for row_index in range(sheet.max_row, 1, -1):
        job_id = str(sheet.cell(row=row_index, column=job_id_column).value or "").strip()
        if job_id in job_ids:
            sheet.delete_rows(row_index, 1)
    workbook.save(tracking_path)


def _schedule_browser_open(*, host: str, port: int) -> None:
    url = _browser_url(host=host, port=port)

    def open_url() -> None:
        import webbrowser

        webbrowser.open(url)

    timer = threading.Timer(1.0, open_url)
    timer.daemon = True
    timer.start()


def _browser_url(*, host: str, port: int) -> str:
    browser_host = "127.0.0.1" if host in {"", "0.0.0.0", "::"} else host
    if ":" in browser_host and not browser_host.startswith("["):
        browser_host = f"[{browser_host}]"
    return f"http://{browser_host}:{port}"


def _open_url_in_chromium(url: str) -> None:
    try:
        subprocess.Popen(  # noqa: S603
            [
                sys.executable,
                "-c",
                _PLAYWRIGHT_CHROMIUM_SCRIPT,
                url,
            ],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            start_new_session=True,
        )
    except OSError:
        _open_url_in_default_browser(url)


def _open_url_in_default_browser(url: str) -> None:
    import webbrowser

    webbrowser.open(url)


_PLAYWRIGHT_CHROMIUM_SCRIPT = r"""
import sys

url = sys.argv[1]

try:
    from playwright.sync_api import sync_playwright
except Exception:
    import webbrowser

    webbrowser.open(url)
    raise SystemExit(0)

try:
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()
        page.goto(url)
        page.wait_for_event("close", timeout=0)
except Exception:
    import webbrowser

    webbrowser.open(url)
"""


INDEX_TEMPLATE = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>LinkedIn Applications</title>
  <style>
    :root {
      color-scheme: light;
      --ink: #202124;
      --muted: #626a73;
      --line: #d9dee5;
      --surface: #ffffff;
      --band: #f4f6f8;
      --accent: #0b6e69;
      --accent-strong: #074f4b;
      --warn: #915c00;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      color: var(--ink);
      background: var(--band);
      font: 14px/1.45 -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }
    header {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      padding: 18px 24px;
      background: var(--surface);
      border-bottom: 1px solid var(--line);
    }
    h1 { margin: 0; font-size: 20px; font-weight: 650; }
    .stats { display: flex; gap: 14px; color: var(--muted); white-space: nowrap; }
    .toolbar {
      display: flex;
      align-items: center;
      gap: 10px;
      padding: 12px 24px;
      background: var(--surface);
      border-bottom: 1px solid var(--line);
      position: sticky;
      top: 0;
      z-index: 2;
    }
    input[type="search"], input[type="date"], textarea, select {
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 8px 9px;
      background: #fff;
      color: var(--ink);
      font: inherit;
    }
    input[type="search"] { max-width: 420px; }
    button, .link-button {
      border: 1px solid var(--accent);
      border-radius: 6px;
      background: var(--accent);
      color: #fff;
      cursor: pointer;
      font: inherit;
      font-weight: 600;
      padding: 8px 10px;
      text-decoration: none;
      white-space: nowrap;
    }
    button:hover, .link-button:hover { background: var(--accent-strong); }
    .sort-button {
      align-items: center;
      background: transparent;
      border: 0;
      color: inherit;
      display: inline-flex;
      font: inherit;
      font-weight: 700;
      gap: 4px;
      padding: 0;
      text-transform: uppercase;
    }
    .sort-button:hover, .sort-button:focus {
      background: transparent;
      color: var(--accent);
      outline: none;
    }
    .sort-indicator {
      color: var(--accent);
      font-size: 11px;
      min-width: 14px;
    }
    .ghost {
      background: #fff;
      color: var(--accent);
    }
    .danger {
      background: #9d2f2f;
      border-color: #9d2f2f;
    }
    .danger:hover {
      background: #7b2424;
      border-color: #7b2424;
    }
    button:disabled {
      cursor: not-allowed;
      opacity: .45;
    }
    .toolbar form { margin: 0; }
    .selected-count { color: var(--muted); white-space: nowrap; }
    main { padding: 20px 24px 32px; }
    .flash {
      margin: 0 0 12px;
      padding: 10px 12px;
      border: 1px solid #c8dfdc;
      border-radius: 6px;
      background: #eef8f6;
      color: var(--accent-strong);
    }
    table {
      width: 100%;
      border-collapse: collapse;
      background: var(--surface);
      border: 1px solid var(--line);
    }
    th, td {
      border-bottom: 1px solid var(--line);
      padding: 10px;
      text-align: left;
      vertical-align: top;
    }
    th {
      color: var(--muted);
      font-size: 12px;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: .02em;
      background: #fafbfc;
      position: sticky;
      top: 51px;
      z-index: 1;
    }
    tr.is-applied { background: #f3faf8; }
    .select-col { width: 42px; text-align: center; }
    .company { min-width: 160px; font-weight: 650; }
    .job { min-width: 260px; }
    .job-id { display: block; color: var(--muted); font-size: 12px; margin-top: 3px; }
    .date-col { min-width: 104px; white-space: nowrap; }
    .actions { display: flex; gap: 8px; flex-wrap: wrap; min-width: 170px; }
    .muted { color: var(--muted); }
    .apply-form {
      display: grid;
      grid-template-columns: 96px 150px minmax(180px, 1fr) auto;
      gap: 8px;
      min-width: 540px;
    }
    textarea { min-height: 38px; resize: vertical; }
    a { color: var(--accent); font-weight: 600; }
    .empty { color: var(--muted); padding: 28px; text-align: center; }
    @media (max-width: 900px) {
      header, .toolbar { align-items: flex-start; flex-direction: column; }
      .stats { flex-wrap: wrap; }
      main { padding: 12px; overflow-x: auto; }
      th { top: 104px; }
    }
  </style>
</head>
<body>
  <header>
    <h1>LinkedIn Applications</h1>
    <div class="stats">
      <span>Total: {{ stats.total }}</span>
      <span>Applied: {{ stats.applied }}</span>
      <span>Pending: {{ stats.pending }}</span>
      <span>N/A: {{ stats.not_applicable }}</span>
    </div>
  </header>
  <div class="toolbar">
    <input id="search" type="search" placeholder="Search company, title, or job id">
    <select id="status-filter" aria-label="Filter status">
      <option value="all">All statuses</option>
      <option value="No">Pending</option>
      <option value="Yes">Applied</option>
      <option value="N/A">N/A</option>
    </select>
    <form method="post" action="/sync">
      <button type="submit" class="ghost">Sync from output</button>
    </form>
    <form id="bulk-delete-form" method="post" action="/applications/delete">
      <button id="delete-selected" type="submit" class="danger" disabled>
        Delete selected
      </button>
    </form>
    <span id="selected-count" class="selected-count">0 selected</span>
  </div>
  <main>
    {% with messages = get_flashed_messages() %}
      {% if messages %}
        {% for message in messages %}
          <p class="flash">{{ message }}</p>
        {% endfor %}
      {% endif %}
    {% endwith %}
    {% if rows %}
      <table id="applications">
        <thead>
          <tr>
            <th class="select-col">
              <input id="select-all" type="checkbox" aria-label="Select visible rows">
            </th>
            <th id="company-header" aria-sort="none">
              <button
                id="company-sort"
                class="sort-button"
                type="button"
                aria-label="Sort by company"
              >
                Company
                <span id="company-sort-indicator" class="sort-indicator" aria-hidden="true">
                  ↑↓
                </span>
              </button>
            </th>
            <th>Job</th>
            <th>Posted</th>
            <th id="matched-header" aria-sort="none">
              <button
                id="matched-sort"
                class="sort-button"
                type="button"
                aria-label="Sort by matched date"
              >
                Matched
                <span id="matched-sort-indicator" class="sort-indicator" aria-hidden="true">
                  ↑↓
                </span>
              </button>
            </th>
            <th>Job Links</th>
            <th>Resume</th>
            <th>Cover Letter</th>
            <th>Application</th>
          </tr>
        </thead>
        <tbody>
          {% for row in rows %}
            <tr class="{{ 'is-applied' if row.applied_to == 'Yes' else '' }}"
                data-status="{{ row.applied_to }}"
                data-company-sort="{{ row.company }}"
                data-matched-sort="{{ row.date_matched or '' }}"
                data-search="{{ (row.company ~ ' ' ~ row.job_title ~ ' ' ~ row.job_id)|lower }}">
              <td class="select-col">
                <input
                  class="row-selector"
                  type="checkbox"
                  name="job_id"
                  value="{{ row.job_id }}"
                  form="bulk-delete-form"
                  aria-label="Select {{ row.company }} {{ row.job_title }}"
                >
              </td>
              <td class="company">{{ row.company }}</td>
              <td class="job">
                {{ row.job_title }}
                <span class="job-id">{{ row.job_id }}</span>
              </td>
              <td class="date-col">
                {{ row.date_posted|display_date if row.date_posted else '' }}
                {% if not row.date_posted %}<span class="muted">-</span>{% endif %}
              </td>
              <td class="date-col">
                {{ row.date_matched|display_date if row.date_matched else '' }}
                {% if not row.date_matched %}<span class="muted">-</span>{% endif %}
              </td>
              <td>
                <div class="actions">
                  <a href="/linkedin/{{ row.job_id }}">LinkedIn</a>
                  <a
                    href="/descriptions/{{ row.job_id }}"
                    target="_blank"
                    rel="noreferrer"
                  >
                    Compare descriptions
                  </a>
                </div>
              </td>
              <td>
                <div class="actions">
                  {% if row.resume_content %}
                    <a href="/resumes/{{ row.job_id }}" target="_blank" rel="noreferrer">
                      Resume
                    </a>
                    <a href="/resumes/{{ row.job_id }}/download">
                      Download
                    </a>
                  {% else %}
                    <span class="muted">Missing</span>
                  {% endif %}
                </div>
              </td>
              <td>
                <div class="actions">
                  {% if row.cover_letter_content %}
                    <a href="/cover-letters/{{ row.job_id }}" target="_blank" rel="noreferrer">
                      Cover Letter
                    </a>
                    <a href="/cover-letters/{{ row.job_id }}/download">
                      Download
                    </a>
                  {% else %}
                    <span class="muted">Missing</span>
                  {% endif %}
                </div>
              </td>
              <td>
                <form class="apply-form" method="post" action="/applications/{{ row.job_id }}">
                  <select name="applied_to" aria-label="Applied status">
                    <option
                      value="No"
                      {{ 'selected' if row.applied_to == 'No' else '' }}
                    >No</option>
                    <option
                      value="Yes"
                      {{ 'selected' if row.applied_to == 'Yes' else '' }}
                    >Yes</option>
                    <option
                      value="N/A"
                      {{ 'selected' if row.applied_to == 'N/A' else '' }}
                    >N/A</option>
                  </select>
                  <input type="date" name="date_applied" value="{{ row.date_applied or '' }}">
                  <textarea name="notes" placeholder="Notes">{{ row.notes }}</textarea>
                  <button type="submit">Save</button>
                </form>
              </td>
            </tr>
          {% endfor %}
        </tbody>
      </table>
    {% else %}
      <div class="empty">No applications imported yet.</div>
    {% endif %}
  </main>
  <script>
    const search = document.querySelector("#search");
    const statusFilter = document.querySelector("#status-filter");
    const selectAll = document.querySelector("#select-all");
    const deleteButton = document.querySelector("#delete-selected");
    const selectedCount = document.querySelector("#selected-count");
    const bulkDeleteForm = document.querySelector("#bulk-delete-form");
    const companySortButton = document.querySelector("#company-sort");
    const companySortIndicator = document.querySelector("#company-sort-indicator");
    const companyHeader = document.querySelector("#company-header");
    const matchedSortButton = document.querySelector("#matched-sort");
    const matchedSortIndicator = document.querySelector("#matched-sort-indicator");
    const matchedHeader = document.querySelector("#matched-header");
    const tableBody = document.querySelector("#applications tbody");
    const rows = [...document.querySelectorAll("#applications tbody tr")];
    const rowSelectors = [...document.querySelectorAll(".row-selector")];
    let companySortDirection = null;
    let matchedSortDirection = null;
    function applyFilters() {
      const term = search.value.trim().toLowerCase();
      const status = statusFilter.value;
      rows.forEach((row) => {
        const matchesText = !term || row.dataset.search.includes(term);
        const matchesStatus = status === "all" || row.dataset.status === status;
        row.hidden = !(matchesText && matchesStatus);
      });
      updateSelectionState();
    }
    function matchedTimestamp(row) {
      const value = row.dataset.matchedSort;
      if (!value) {
        return null;
      }
      const parsed = Date.parse(value);
      return Number.isNaN(parsed) ? null : parsed;
    }
    function resetSortIndicator(header, indicator) {
      if (header) {
        header.setAttribute("aria-sort", "none");
      }
      if (indicator) {
        indicator.textContent = "↑↓";
      }
    }
    function sortRowsByCompany() {
      if (!tableBody) {
        return;
      }
      companySortDirection = companySortDirection === "asc" ? "desc" : "asc";
      matchedSortDirection = null;
      resetSortIndicator(matchedHeader, matchedSortIndicator);
      const direction = companySortDirection === "asc" ? 1 : -1;
      const originalIndex = new Map(rows.map((row, index) => [row, index]));
      const sortedRows = [...rows].sort((left, right) => {
        const leftValue = (left.dataset.companySort || "").trim();
        const rightValue = (right.dataset.companySort || "").trim();
        const comparison = leftValue.localeCompare(
          rightValue,
          undefined,
          { sensitivity: "base" },
        );
        if (comparison === 0) {
          return originalIndex.get(left) - originalIndex.get(right);
        }
        return comparison * direction;
      });
      sortedRows.forEach((row) => tableBody.appendChild(row));
      if (companyHeader) {
        companyHeader.setAttribute(
          "aria-sort",
          companySortDirection === "asc" ? "ascending" : "descending",
        );
      }
      companySortIndicator.textContent = companySortDirection === "asc" ? "↑" : "↓";
      applyFilters();
    }
    function sortRowsByMatched() {
      if (!tableBody) {
        return;
      }
      matchedSortDirection = matchedSortDirection === "asc" ? "desc" : "asc";
      companySortDirection = null;
      resetSortIndicator(companyHeader, companySortIndicator);
      const direction = matchedSortDirection === "asc" ? 1 : -1;
      const originalIndex = new Map(rows.map((row, index) => [row, index]));
      const sortedRows = [...rows].sort((left, right) => {
        const leftValue = matchedTimestamp(left);
        const rightValue = matchedTimestamp(right);
        if (leftValue === null && rightValue === null) {
          return originalIndex.get(left) - originalIndex.get(right);
        }
        if (leftValue === null) {
          return 1;
        }
        if (rightValue === null) {
          return -1;
        }
        if (leftValue === rightValue) {
          return originalIndex.get(left) - originalIndex.get(right);
        }
        return (leftValue - rightValue) * direction;
      });
      sortedRows.forEach((row) => tableBody.appendChild(row));
      if (matchedHeader) {
        matchedHeader.setAttribute(
          "aria-sort",
          matchedSortDirection === "asc" ? "ascending" : "descending",
        );
      }
      matchedSortIndicator.textContent = matchedSortDirection === "asc" ? "↑" : "↓";
      applyFilters();
    }
    function visibleSelectors() {
      return rowSelectors.filter((checkbox) => !checkbox.closest("tr").hidden);
    }
    function updateSelectionState() {
      const selected = rowSelectors.filter((checkbox) => checkbox.checked);
      const visible = visibleSelectors();
      deleteButton.disabled = selected.length === 0;
      selectedCount.textContent = `${selected.length} selected`;
      if (!selectAll) {
        return;
      }
      selectAll.checked = visible.length > 0 && visible.every((checkbox) => checkbox.checked);
      selectAll.indeterminate = !selectAll.checked
        && visible.some((checkbox) => checkbox.checked);
    }
    search.addEventListener("input", applyFilters);
    statusFilter.addEventListener("change", applyFilters);
    if (companySortButton) {
      companySortButton.addEventListener("click", sortRowsByCompany);
    }
    if (matchedSortButton) {
      matchedSortButton.addEventListener("click", sortRowsByMatched);
    }
    if (selectAll) {
      selectAll.addEventListener("change", () => {
        visibleSelectors().forEach((checkbox) => {
          checkbox.checked = selectAll.checked;
        });
        updateSelectionState();
      });
      rowSelectors.forEach((checkbox) => {
        checkbox.addEventListener("change", updateSelectionState);
      });
    }
    bulkDeleteForm.addEventListener("submit", (event) => {
      const selected = rowSelectors.filter((checkbox) => checkbox.checked);
      if (selected.length === 0) {
        event.preventDefault();
        return;
      }
      if (!confirm(`Delete ${selected.length} selected application row(s)?`)) {
        event.preventDefault();
      }
    });
  </script>
</body>
</html>
"""


DESCRIPTION_COMPARE_TEMPLATE = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Description Compare - {{ row.company }} - {{ row.job_title }}</title>
  <style>
    :root {
      color-scheme: light;
      --ink: #202124;
      --muted: #626a73;
      --line: #d9dee5;
      --surface: #ffffff;
      --band: #f4f6f8;
      --accent: #0b6e69;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      color: var(--ink);
      background: var(--band);
      font: 14px/1.45 -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }
    header {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      padding: 18px 24px;
      background: var(--surface);
      border-bottom: 1px solid var(--line);
    }
    h1 { margin: 0; font-size: 20px; font-weight: 650; }
    .meta { color: var(--muted); margin-top: 4px; }
    a { color: var(--accent); font-weight: 650; }
    main {
      display: grid;
      grid-template-columns: minmax(0, 1fr) minmax(0, 1fr);
      gap: 16px;
      padding: 16px 24px 24px;
    }
    section {
      min-width: 0;
      background: var(--surface);
      border: 1px solid var(--line);
    }
    h2 {
      margin: 0;
      padding: 12px 14px;
      border-bottom: 1px solid var(--line);
      font-size: 14px;
      font-weight: 700;
    }
    pre {
      min-height: calc(100vh - 160px);
      max-height: calc(100vh - 160px);
      margin: 0;
      overflow: auto;
      padding: 14px;
      white-space: pre-wrap;
      word-break: break-word;
      background: #fff;
      color: var(--ink);
      font: 13px/1.5 ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
    }
    @media (max-width: 900px) {
      header {
        align-items: flex-start;
        flex-direction: column;
      }
      main {
        grid-template-columns: 1fr;
        padding: 12px;
      }
      pre {
        min-height: 48vh;
        max-height: 48vh;
      }
    }
  </style>
</head>
<body>
  <header>
    <div>
      <h1>{{ row.job_title }}</h1>
      <div class="meta">{{ row.company }} · {{ row.job_id }}</div>
    </div>
    <a href="/" target="_self">Applications</a>
  </header>
  <main>
    <section>
      <h2>Parsed Job Description</h2>
      <pre>{{ row.job_description or "No parsed job description is stored." }}</pre>
    </section>
    <section>
      <h2>Prompt Job Description</h2>
      <pre>{{ row.prompt_job_description or "No prompt job description is stored." }}</pre>
    </section>
  </main>
</body>
</html>
"""


if __name__ == "__main__":
    main()
