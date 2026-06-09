from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

from linkedin_career_mcp.config import Settings
from linkedin_career_mcp.models import JobDetails, JobPosting, JobSearchQuery, JobSearchResult
from linkedin_career_mcp.webapp import (
    DEFAULT_DATABASE,
    connect_database,
    fetch_existing_resume_job_ids,
    upsert_application_artifact,
)
from linkedin_career_mcp.workflows import matching
from linkedin_career_mcp.workflows.matching import (
    DEFAULT_SUPPLEMENTAL_SEARCH_KEYWORDS,
    CompanyBlacklist,
    MatchingJobsWorkflow,
    MatchingJobsWorkflowResult,
    TailoredResumeArtifact,
    _coerce_search_query,
    _expand_remote_and_hybrid_queries,
    _normalize_regeneration_job_ids,
    _supplement_search_queries,
)


class FakeOllama:
    def __init__(self) -> None:
        self.text_prompts: list[str] = []
        self.json_prompts: list[str] = []

    async def generate_json(self, prompt: str) -> dict[str, object]:
        self.json_prompts.append(prompt)
        if "cover_letter_sections" in prompt:
            return {
                "opening_alignment": (
                    "the AI platform engineering and automation capabilities you are looking for"
                ),
                "oracle_alignment": (
                    "I have owned Oracle automation components and built observability pipelines "
                    "that map to this role's distributed systems needs."
                ),
                "prior_experience_alignment": (
                    "My pre-Oracle roles add healthcare automation, React development, Azure "
                    "migration, and Django platform experience relevant to this position."
                ),
            }
        if "core_technical_skills" in prompt:
            return {
                "core_technical_skills": [
                    {
                        "category": "Languages & Frameworks",
                        "skills": ["Python", "JavaScript", "Node.js", "React.js"],
                    },
                    {
                        "category": "Distributed Systems & Cloud",
                        "skills": ["AWS", "Azure", "OpenSearch"],
                    },
                    {
                        "category": "Platform & API Engineering",
                        "skills": ["RESTful APIs", "Systems Architecture", "Microservices"],
                    },
                    {
                        "category": "Automation & IaC",
                        "skills": ["Terraform", "Ansible", "Jenkins"],
                    },
                    {
                        "category": "Data & Observability",
                        "skills": ["Data Pipelines", "Observability Dashboards"],
                    },
                    {
                        "category": "Security & Compliance",
                        "skills": ["Secure Coding Practices", "RBAC"],
                    },
                    {
                        "category": "AI Tools",
                        "skills": ["Codex", "Oracle Code Assist (OCA)", "Cline", "OpenRouter"],
                    },
                ],
                "prior_experience": [],
            }
        return {
            "queries": [
                {
                    "keywords": "agentic ai platform engineer",
                    "location": "United States",
                    "date_posted": "past_week",
                    "job_type": "full_time",
                    "sort_by": "recent",
                    "limit": 5,
                }
            ]
        }

    async def generate_text(self, prompt: str) -> str:
        self.text_prompts.append(prompt)
        return (
            "**Oracle | Remote / International Datacenters**\n"
            "**Senior Technical Lead - Cloud Automation Engineer | Feb 2022 - Present**\n"
            "- **Platform Component Ownership:** Built agentic AI systems and MCP tools."
        )


class FlakyCoverLetterOllama(FakeOllama):
    def __init__(self) -> None:
        super().__init__()
        self.cover_letter_attempts = 0

    async def generate_json(self, prompt: str) -> dict[str, object]:
        if "cover_letter_sections" in prompt:
            self.cover_letter_attempts += 1
            if self.cover_letter_attempts == 1:
                raise matching.WorkflowError("temporary cover-letter failure")
        return await super().generate_json(prompt)


class FakeJobService:
    def __init__(self) -> None:
        self.queries: list[JobSearchQuery] = []

    async def search(self, query: JobSearchQuery) -> JobSearchResult:
        self.queries.append(query)
        return JobSearchResult(
            query=query,
            count=2,
            jobs=[
                JobPosting(
                    job_id="111",
                    title="Agentic AI Engineer",
                    company="Acme AI",
                    job_url="https://www.linkedin.com/jobs/view/111",
                ),
                JobPosting(
                    job_id="222",
                    title="Defense AI Engineer",
                    company="Raytheon Technologies",
                    job_url="https://www.linkedin.com/jobs/view/222",
                ),
            ],
            provider="fake",
        )

    async def get_details(self, job_id_or_url: str) -> JobDetails:
        job_id = job_id_or_url.rstrip("/").rsplit("/", 1)[-1]
        if job_id == "222":
            return JobDetails(
                job_id="222",
                title="Defense AI Engineer",
                company="Raytheon Technologies",
                job_url="https://www.linkedin.com/jobs/view/222",
                description="Defense AI role.",
            )
        return JobDetails(
            job_id="111",
            title="Agentic AI Engineer",
            company="Acme AI",
            job_url="https://www.linkedin.com/jobs/view/111",
            listed_at="2026-06-07T09:30:00Z",
            posted_text="1 day ago",
            workplace_type="Remote",
            description="Build local AI workflows and MCP integrations.",
        )


class SkillRepairJobService:
    def __init__(self) -> None:
        self.queries: list[JobSearchQuery] = []

    async def search(self, query: JobSearchQuery) -> JobSearchResult:
        self.queries.append(query)
        return JobSearchResult(
            query=query,
            count=1,
            jobs=[
                JobPosting(
                    job_id="777",
                    title="Platform Automation Engineer",
                    company="Skillful Co",
                    job_url="https://www.linkedin.com/jobs/view/777",
                ),
            ],
            provider="fake",
        )

    async def get_details(self, job_id_or_url: str) -> JobDetails:
        return JobDetails(
            job_id="777",
            title="Platform Automation Engineer",
            company="Skillful Co",
            job_url="https://www.linkedin.com/jobs/view/777",
            workplace_type="Remote",
            description=(
                "Required experience with Kubernetes, TypeScript, and GitHub Actions "
                "for platform automation and cloud infrastructure."
            ),
        )


class ExistingThenNewJobService:
    def __init__(self) -> None:
        self.queries: list[JobSearchQuery] = []
        self.detail_requests: list[str] = []

    async def search(self, query: JobSearchQuery) -> JobSearchResult:
        self.queries.append(query)
        if len(self.queries) == 1:
            jobs = [
                JobPosting(
                    job_id="111",
                    title="Existing Engineer",
                    company="Existing Co",
                    job_url="https://www.linkedin.com/jobs/view/111",
                )
            ]
        else:
            jobs = [
                JobPosting(
                    job_id="333",
                    title="Fresh Platform Engineer",
                    company="Fresh Co",
                    job_url="https://www.linkedin.com/jobs/view/333",
                )
            ]
        return JobSearchResult(query=query, count=len(jobs), jobs=jobs, provider="fake")

    async def get_details(self, job_id_or_url: str) -> JobDetails:
        self.detail_requests.append(job_id_or_url)
        job_id = job_id_or_url.rstrip("/").rsplit("/", 1)[-1]
        return JobDetails(
            job_id=job_id,
            title="Fresh Platform Engineer",
            company="Fresh Co",
            job_url=f"https://www.linkedin.com/jobs/view/{job_id}",
            description="Fresh role that should count toward max_jobs.",
        )


class OnSiteLeakJobService:
    def __init__(self) -> None:
        self.queries: list[JobSearchQuery] = []

    async def search(self, query: JobSearchQuery) -> JobSearchResult:
        self.queries.append(query)
        return JobSearchResult(
            query=query,
            count=2,
            jobs=[
                JobPosting(
                    job_id="444",
                    title="Onsite Platform Engineer",
                    company="Office Co",
                    location="Austin, TX (On-site)",
                    workplace_type="On-site",
                    job_url="https://www.linkedin.com/jobs/view/444",
                ),
                JobPosting(
                    job_id="555",
                    title="Remote Platform Engineer",
                    company="Remote Co",
                    location="United States",
                    workplace_type="Remote",
                    job_url="https://www.linkedin.com/jobs/view/555",
                ),
            ],
            provider="fake",
        )

    async def get_details(self, job_id_or_url: str) -> JobDetails:
        job_id = job_id_or_url.rstrip("/").rsplit("/", 1)[-1]
        if job_id == "444":
            return JobDetails(
                job_id="444",
                title="Onsite Platform Engineer",
                company="Office Co",
                job_url="https://www.linkedin.com/jobs/view/444",
                description="On-site role that leaked through LinkedIn search.",
            )
        return JobDetails(
            job_id="555",
            title="Remote Platform Engineer",
            company="Remote Co",
            job_url="https://www.linkedin.com/jobs/view/555",
            description="Remote role that should be accepted.",
        )


class ExperienceLeakJobService:
    def __init__(self) -> None:
        self.queries: list[JobSearchQuery] = []

    async def search(self, query: JobSearchQuery) -> JobSearchResult:
        self.queries.append(query)
        return JobSearchResult(
            query=query,
            count=3,
            jobs=[
                JobPosting(
                    job_id="666",
                    title="Software Engineering Intern",
                    company="Intern Co",
                    workplace_type="Remote",
                    job_url="https://www.linkedin.com/jobs/view/666",
                ),
                JobPosting(
                    job_id="777",
                    title="Entry Level Software Engineer",
                    company="Entry Co",
                    workplace_type="Remote",
                    job_url="https://www.linkedin.com/jobs/view/777",
                ),
                JobPosting(
                    job_id="888",
                    title="Senior Platform Engineer",
                    company="Senior Co",
                    workplace_type="Remote",
                    job_url="https://www.linkedin.com/jobs/view/888",
                ),
            ],
            provider="fake",
        )

    async def get_details(self, job_id_or_url: str) -> JobDetails:
        job_id = job_id_or_url.rstrip("/").rsplit("/", 1)[-1]
        if job_id == "666":
            return JobDetails(
                job_id="666",
                title="Software Engineering Intern",
                company="Intern Co",
                job_url="https://www.linkedin.com/jobs/view/666",
                workplace_type="Remote",
                seniority_level="Internship",
                description="Internship role that should be skipped.",
            )
        if job_id == "777":
            return JobDetails(
                job_id="777",
                title="Entry Level Software Engineer",
                company="Entry Co",
                job_url="https://www.linkedin.com/jobs/view/777",
                workplace_type="Remote",
                seniority_level="Entry level",
                description="Entry-level role that should be skipped.",
            )
        return JobDetails(
            job_id="888",
            title="Senior Platform Engineer",
            company="Senior Co",
            job_url="https://www.linkedin.com/jobs/view/888",
            workplace_type="Remote",
            seniority_level="Mid-Senior level",
            description="Senior platform role that should be accepted.",
        )


def test_company_blacklist_matches_globs_case_insensitively(tmp_path: Path):
    blacklist_path = tmp_path / ".blacklist"
    blacklist_path.write_text("# comment\nRaytheon*\n", encoding="utf-8")

    blacklist = CompanyBlacklist.from_file(blacklist_path)

    assert blacklist.matches("Raytheon Technologies")
    assert blacklist.matches("raytheon")
    assert not blacklist.matches("Acme AI")


def test_search_queries_are_supplemented_with_trending_keyword_fallbacks():
    model_query = JobSearchQuery(
        keywords="Senior Platform Engineer AWS Azure Terraform",
        location="United States",
        date_posted="past_month",
        job_type="full_time",
        workplace_type="remote",
        experience_level="mid_senior",
        sort_by="recent",
        limit=5,
    )

    supplemented = _supplement_search_queries(
        [model_query],
        location="United States",
        date_posted="past_month",
        limit_per_query=5,
    )
    expanded = _expand_remote_and_hybrid_queries(supplemented, max_queries=8)

    keywords = [query.keywords for query in expanded]
    assert keywords[:2] == [model_query.keywords, model_query.keywords]
    assert DEFAULT_SUPPLEMENTAL_SEARCH_KEYWORDS[0] in keywords
    assert any("AI" in keyword or "LLM" in keyword or "agentic" in keyword for keyword in keywords)
    assert {query.workplace_type for query in expanded[:2]} == {"remote", "hybrid"}


def test_coerce_search_query_drops_internship_and_entry_level_filters():
    for experience_level in ("internship", "entry_level", "Entry level"):
        query = _coerce_search_query(
            {
                "keywords": "software engineer",
                "experience_level": experience_level,
            },
            location="United States",
            date_posted="past_week",
            limit_per_query=5,
        )

        assert query.experience_level is None


def test_normalize_regeneration_job_ids_accepts_csv_and_positional_ids():
    assert _normalize_regeneration_job_ids(["111"]) == ["111"]
    assert _normalize_regeneration_job_ids(["111,222", "333"]) == ["111", "222", "333"]
    assert _normalize_regeneration_job_ids(["111, 222", "222"]) == ["111", "222"]
    assert _normalize_regeneration_job_ids(["all"]) is None


async def test_matching_workflow_writes_resume_and_tracking(tmp_path: Path):
    profile_dir = tmp_path / "profile"
    profile_dir.mkdir()
    (profile_dir / "MP-RESUME-AGENTIC.txt").write_text(
        "Resume: built MCP servers and local LLM workflows.",
        encoding="utf-8",
    )
    blacklist_path = tmp_path / ".blacklist"
    blacklist_path.write_text("Raytheon*\n", encoding="utf-8")
    output_dir = tmp_path / "output"
    service = FakeJobService()
    workflow = MatchingJobsWorkflow(service=service, ollama=FakeOllama())

    result = await workflow.run(
        profile_dir=profile_dir,
        blacklist_path=blacklist_path,
        output_dir=output_dir,
        source_resume_name="MP-RESUME-AGENTIC.txt",
        limit_per_query=5,
        max_queries=2,
        max_jobs=5,
    )

    assert {query.workplace_type for query in service.queries} == {"remote", "hybrid"}
    assert result.resumes_created == 1
    assert result.cover_letters_created == 1
    assert result.artifacts[0].company == "Acme AI"
    assert "Raytheon Technologies - Defense AI Engineer" in result.skipped_blacklisted

    resume_path = Path(result.artifacts[0].resume_path)
    assert resume_path.exists()
    assert resume_path.name == "mp_resume_agentic_ai_engineer.pdf"
    resume_text = "\n".join(page.extract_text() or "" for page in PdfReader(resume_path).pages)
    assert "Max Perkhounkov" in resume_text
    assert "custom tailored for every job position" in resume_text
    assert "Education & Certifications" in resume_text
    assert "Oracle Cloud Infrastructure AI Foundations Associate" in resume_text
    assert "AI Tools" in resume_text
    assert "OpenRouter" in resume_text
    assert "Error Budgets" not in resume_text
    assert "**" not in resume_text

    cover_letter_path = Path(result.artifacts[0].cover_letter_path or "")
    assert cover_letter_path.exists()
    assert cover_letter_path.name == "mp_cover_letter_agentic_ai_engineer.pdf"
    cover_letter_text = "\n".join(
        page.extract_text() or "" for page in PdfReader(cover_letter_path).pages
    )
    assert "Dear Hiring Manager" in cover_letter_text
    assert "Agentic AI Engineer at Acme AI" in cover_letter_text
    assert "linkedin-career-mcp" in cover_letter_text
    assert "OpenRouter and DeepSeek" in cover_letter_text

    database_path = output_dir / DEFAULT_DATABASE
    with connect_database(database_path) as connection:
        row = connection.execute(
            """
            SELECT job_description, prompt_job_description, cover_letter_content,
                   source_cover_letter_path, date_matched, date_posted
            FROM applications
            WHERE job_id = ?
            """,
            ("111",),
        ).fetchone()
    assert row["job_description"] == "Build local AI workflows and MCP integrations."
    assert row["prompt_job_description"] == "Build local AI workflows and MCP integrations."
    assert row["cover_letter_content"] is not None
    assert row["source_cover_letter_path"] == str(cover_letter_path)
    assert row["date_matched"]
    assert row["date_posted"] == "2026-06-07"

    workbook_path = output_dir / "tracking/read_applications/linkedin_applications.xlsx"
    assert workbook_path.exists()
    workbook = load_workbook(workbook_path)
    sheet = workbook.active
    assert sheet.cell(row=2, column=1).value == "111"
    assert sheet.cell(row=2, column=4).hyperlink.target == "https://www.linkedin.com/jobs/view/111"
    assert sheet.cell(row=2, column=6).value == str(cover_letter_path)
    assert sheet.cell(row=2, column=6).hyperlink.target == cover_letter_path.resolve().as_uri()
    assert sheet.cell(row=2, column=7).value == "No"
    assert sheet.cell(row=2, column=8).value is None


async def test_matching_workflow_repairs_ats_missing_skills_from_profile_skills(
    tmp_path: Path,
):
    profile_dir = tmp_path / "profile"
    profile_dir.mkdir()
    (profile_dir / "MP-RESUME-AGENTIC.txt").write_text(
        "Resume: built platform automation workflows.",
        encoding="utf-8",
    )
    (profile_dir / "skills.md").write_text(
        "\n".join(
            [
                "* **Languages & Frameworks:** TypeScript",
                "* **Distributed Systems & Cloud:** Kubernetes",
                "* **Automation & IaC (DevOps):** GitHub Actions",
                "* **Data & Observability:** Error Budgets",
                "* **Networking hardware:** Palo-Alto",
            ]
        ),
        encoding="utf-8",
    )
    blacklist_path = tmp_path / ".blacklist"
    blacklist_path.write_text("", encoding="utf-8")
    output_dir = tmp_path / "output"
    progress_messages: list[str] = []
    workflow = MatchingJobsWorkflow(
        service=SkillRepairJobService(),
        ollama=FakeOllama(),
    )

    result = await workflow.run(
        profile_dir=profile_dir,
        blacklist_path=blacklist_path,
        output_dir=output_dir,
        source_resume_name="MP-RESUME-AGENTIC.txt",
        limit_per_query=5,
        max_queries=1,
        max_jobs=1,
        artifact_mode="resumes-only",
        progress_callback=progress_messages.append,
    )

    assert result.errors == []
    resume_path = Path(result.artifacts[0].resume_path)
    resume_text = "\n".join(page.extract_text() or "" for page in PdfReader(resume_path).pages)
    assert "TypeScript" in resume_text
    assert "Kubernetes" in resume_text
    assert "GitHub Actions" in resume_text
    assert "Automation & IaC: Terraform, Ansible, Jenkins, GitHub Actions" in resume_text
    assert "Error Budgets" not in resume_text
    assert "Palo-Alto" not in resume_text
    assert any("ATS skill repair added" in message for message in progress_messages)

    database_path = output_dir / DEFAULT_DATABASE
    with connect_database(database_path) as connection:
        row = connection.execute(
            """
            SELECT ats_score, ats_missing_terms
            FROM applications
            WHERE job_id = ?
            """,
            ("777",),
        ).fetchone()
    assert row["ats_score"] is not None
    missing_terms = (row["ats_missing_terms"] or "").casefold()
    assert "kubernetes" not in missing_terms
    assert "typescript" not in missing_terms
    assert "github actions" not in missing_terms


async def test_matching_workflow_skips_explicit_on_site_linkedin_leaks(tmp_path: Path):
    profile_dir = tmp_path / "profile"
    profile_dir.mkdir()
    (profile_dir / "MP-RESUME-AGENTIC.txt").write_text(
        "Resume: built MCP servers and local LLM workflows.",
        encoding="utf-8",
    )
    blacklist_path = tmp_path / ".blacklist"
    blacklist_path.write_text("", encoding="utf-8")
    output_dir = tmp_path / "output"
    service = OnSiteLeakJobService()
    workflow = MatchingJobsWorkflow(service=service, ollama=FakeOllama())

    result = await workflow.run(
        profile_dir=profile_dir,
        blacklist_path=blacklist_path,
        output_dir=output_dir,
        source_resume_name="MP-RESUME-AGENTIC.txt",
        limit_per_query=5,
        max_queries=2,
        max_jobs=1,
    )

    assert result.jobs_found == 1
    assert result.artifacts[0].job_id == "555"
    assert "Office Co - Onsite Platform Engineer" in result.skipped_workplace_type


async def test_matching_workflow_skips_internship_and_entry_level_jobs(tmp_path: Path):
    profile_dir = tmp_path / "profile"
    profile_dir.mkdir()
    (profile_dir / "MP-RESUME-AGENTIC.txt").write_text(
        "Resume: built platform automation systems.",
        encoding="utf-8",
    )
    blacklist_path = tmp_path / ".blacklist"
    blacklist_path.write_text("", encoding="utf-8")
    output_dir = tmp_path / "output"
    workflow = MatchingJobsWorkflow(
        service=ExperienceLeakJobService(),
        ollama=FakeOllama(),
    )

    result = await workflow.run(
        profile_dir=profile_dir,
        blacklist_path=blacklist_path,
        output_dir=output_dir,
        source_resume_name="MP-RESUME-AGENTIC.txt",
        limit_per_query=5,
        max_queries=1,
        max_jobs=1,
        artifact_mode="resumes-only",
    )

    assert result.jobs_found == 1
    assert result.artifacts[0].job_id == "888"
    assert "Intern Co - Software Engineering Intern" in result.skipped_experience_level
    assert "Entry Co - Entry Level Software Engineer" in result.skipped_experience_level

    database_path = output_dir / DEFAULT_DATABASE
    with connect_database(database_path) as connection:
        row = connection.execute(
            """
            SELECT experience_level
            FROM applications
            WHERE job_id = ?
            """,
            ("888",),
        ).fetchone()
    assert row["experience_level"] == "Mid-Senior level"


async def test_matching_workflow_skips_existing_database_jobs_without_counting_them(
    tmp_path: Path,
):
    profile_dir = tmp_path / "profile"
    profile_dir.mkdir()
    (profile_dir / "MP-RESUME-AGENTIC.txt").write_text(
        "Resume: built MCP servers and local LLM workflows.",
        encoding="utf-8",
    )
    blacklist_path = tmp_path / ".blacklist"
    blacklist_path.write_text("", encoding="utf-8")
    output_dir = tmp_path / "output"
    existing_resume = output_dir / "resumes/Existing_Co/111_existing_engineer/resume.pdf"
    existing_resume.parent.mkdir(parents=True)
    existing_resume.write_bytes(b"%PDF-1.4 existing")
    database_path = output_dir / DEFAULT_DATABASE
    upsert_application_artifact(
        database_path=database_path,
        job_id="111",
        company="Existing Co",
        job_title="Existing Engineer",
        linkedin_url="https://www.linkedin.com/jobs/view/111",
        resume_path=existing_resume,
    )
    service = ExistingThenNewJobService()
    workflow = MatchingJobsWorkflow(service=service, ollama=FakeOllama())

    result = await workflow.run(
        profile_dir=profile_dir,
        blacklist_path=blacklist_path,
        output_dir=output_dir,
        source_resume_name="MP-RESUME-AGENTIC.txt",
        limit_per_query=5,
        max_queries=2,
        max_jobs=1,
    )

    assert len(service.queries) == 2
    assert "111" in service.queries[0].exclude_job_ids
    assert service.detail_requests == ["https://www.linkedin.com/jobs/view/333"]
    assert result.jobs_found == 1
    assert result.artifacts[0].job_id == "333"
    assert "Existing Co - Existing Engineer" in result.skipped_existing
    assert fetch_existing_resume_job_ids(database_path) == {"111", "333"}


async def test_regenerate_resumes_uses_database_jobs_and_fetches_missing_descriptions(
    tmp_path: Path,
):
    profile_dir = tmp_path / "profile"
    profile_dir.mkdir()
    (profile_dir / "MP-RESUME-AGENTIC.txt").write_text(
        "Resume: built MCP servers and local LLM workflows.",
        encoding="utf-8",
    )
    output_dir = tmp_path / "output"
    database_path = output_dir / DEFAULT_DATABASE
    stored_resume = output_dir / "resumes/Existing_Co/111_existing_engineer/resume.pdf"
    stored_resume.parent.mkdir(parents=True)
    stored_resume.write_bytes(b"%PDF-1.4 existing")
    missing_description_resume = output_dir / "resumes/Fresh_Co/333_fresh_platform/resume.pdf"
    missing_description_resume.parent.mkdir(parents=True)
    missing_description_resume.write_bytes(b"%PDF-1.4 existing")
    upsert_application_artifact(
        database_path=database_path,
        job_id="111",
        company="Existing Co",
        job_title="Existing Engineer",
        linkedin_url="https://www.linkedin.com/jobs/view/111",
        resume_path=stored_resume,
        job_description="Raw stored JOD with AI Experience.",
        prompt_job_description="Over-trimmed stored prompt JOD.",
    )
    upsert_application_artifact(
        database_path=database_path,
        job_id="333",
        company="Fresh Co",
        job_title="Fresh Platform Engineer",
        linkedin_url="https://www.linkedin.com/jobs/view/333",
        resume_path=missing_description_resume,
    )
    service = ExistingThenNewJobService()
    ollama = FakeOllama()
    workflow = MatchingJobsWorkflow(service=service, ollama=ollama)

    result = await workflow.regenerate_resumes(
        profile_dir=profile_dir,
        output_dir=output_dir,
        source_resume_name="MP-RESUME-AGENTIC.txt",
        job_ids=["111", "333"],
        linkedin_delay_seconds=0,
    )

    assert result.jobs_found == 2
    assert result.resumes_created == 2
    assert [artifact.job_id for artifact in result.artifacts] == ["111", "333"]
    assert service.detail_requests == ["https://www.linkedin.com/jobs/view/333"]
    assert any("Raw stored JOD with AI Experience." in prompt for prompt in ollama.text_prompts)
    assert not any("Over-trimmed stored prompt JOD." in prompt for prompt in ollama.text_prompts)
    with connect_database(database_path) as connection:
        rows = connection.execute(
            """
            SELECT job_id, job_description, prompt_job_description
            FROM applications
            ORDER BY job_id
            """
        ).fetchall()
    row_by_job_id = {row["job_id"]: row for row in rows}
    assert row_by_job_id["111"]["job_description"] == "Raw stored JOD with AI Experience."
    assert row_by_job_id["111"]["prompt_job_description"] == "Raw stored JOD with AI Experience."
    assert row_by_job_id["333"]["job_description"] == (
        "Fresh role that should count toward max_jobs."
    )
    assert row_by_job_id["333"]["prompt_job_description"] == (
        "Fresh role that should count toward max_jobs."
    )


async def test_regenerate_cover_letters_uses_database_jobs_without_regenerating_resumes(
    tmp_path: Path,
):
    profile_dir = tmp_path / "profile"
    profile_dir.mkdir()
    (profile_dir / "MP-RESUME-AGENTIC.txt").write_text(
        "Resume: built MCP servers and local LLM workflows.",
        encoding="utf-8",
    )
    output_dir = tmp_path / "output"
    database_path = output_dir / DEFAULT_DATABASE
    stored_resume = output_dir / "resumes/Existing_Co/111_existing_engineer/resume.pdf"
    stored_resume.parent.mkdir(parents=True)
    stored_resume.write_bytes(b"%PDF-1.4 existing")
    upsert_application_artifact(
        database_path=database_path,
        job_id="111",
        company="Existing Co",
        job_title="Existing Engineer",
        linkedin_url="https://www.linkedin.com/jobs/view/111",
        resume_path=stored_resume,
        job_description="Raw stored JOD with AI Experience.",
        prompt_job_description="Raw stored JOD with AI Experience.",
    )
    service = ExistingThenNewJobService()
    ollama = FakeOllama()
    workflow = MatchingJobsWorkflow(service=service, ollama=ollama)

    result = await workflow.regenerate_resumes(
        profile_dir=profile_dir,
        output_dir=output_dir,
        source_resume_name="MP-RESUME-AGENTIC.txt",
        job_ids=["111"],
        linkedin_delay_seconds=0,
        artifact_mode="cover-letters-only",
    )

    assert result.jobs_found == 1
    assert result.resumes_created == 0
    assert result.cover_letters_created == 1
    assert result.artifacts[0].artifact_kind == "cover_letter"
    assert result.artifacts[0].resume_path == ""
    cover_letter_path = Path(result.artifacts[0].cover_letter_path or "")
    assert cover_letter_path.exists()
    assert service.detail_requests == []
    assert not ollama.text_prompts
    assert any("Raw stored JOD with AI Experience." in prompt for prompt in ollama.json_prompts)
    with connect_database(database_path) as connection:
        row = connection.execute(
            """
            SELECT resume_content, cover_letter_content, source_resume_path,
                   source_cover_letter_path
            FROM applications
            WHERE job_id = ?
            """,
            ("111",),
        ).fetchone()
    assert row["resume_content"] == b"%PDF-1.4 existing"
    assert row["cover_letter_content"] is not None
    assert row["source_resume_path"] == str(stored_resume)
    assert row["source_cover_letter_path"] == str(cover_letter_path)


async def test_regenerate_cover_letters_retries_failed_generation_and_reports_progress(
    tmp_path: Path,
):
    profile_dir = tmp_path / "profile"
    profile_dir.mkdir()
    (profile_dir / "MP-RESUME-AGENTIC.txt").write_text(
        "Resume: built MCP servers and local LLM workflows.",
        encoding="utf-8",
    )
    output_dir = tmp_path / "output"
    database_path = output_dir / DEFAULT_DATABASE
    stored_resume = output_dir / "resumes/Existing_Co/111_existing_engineer/resume.pdf"
    stored_resume.parent.mkdir(parents=True)
    stored_resume.write_bytes(b"%PDF-1.4 existing")
    stored_cover_letter = (
        output_dir / "cover_letters/Existing_Co/111_existing_engineer/cover_letter.pdf"
    )
    stored_cover_letter.parent.mkdir(parents=True)
    stored_cover_letter.write_bytes(b"%PDF-1.4 stale cover")
    upsert_application_artifact(
        database_path=database_path,
        job_id="111",
        company="Existing Co",
        job_title="Existing Engineer",
        linkedin_url="https://www.linkedin.com/jobs/view/111",
        resume_path=stored_resume,
        cover_letter_path=stored_cover_letter,
        job_description="Raw stored JOD with AI Experience.",
        prompt_job_description="Raw stored JOD with AI Experience.",
    )
    service = ExistingThenNewJobService()
    ollama = FlakyCoverLetterOllama()
    workflow = MatchingJobsWorkflow(service=service, ollama=ollama)
    progress_messages: list[str] = []

    result = await workflow.regenerate_resumes(
        profile_dir=profile_dir,
        output_dir=output_dir,
        source_resume_name="MP-RESUME-AGENTIC.txt",
        job_ids=["111"],
        linkedin_delay_seconds=0,
        artifact_mode="cover-letters-only",
        cover_letter_retry_attempts=1,
        progress_callback=progress_messages.append,
    )

    assert result.jobs_found == 1
    assert result.cover_letters_created == 1
    assert result.errors == ["111: temporary cover-letter failure"]
    assert len(result.cover_letter_retries) == 1
    assert result.cover_letter_retries[0].status == "created"
    assert result.artifact_audit.total_jobs == 1
    assert result.artifact_audit.with_cover_letters == 1
    assert result.artifact_audit.missing_artifacts == []
    assert any(
        "Working on job 1/1 - job title: 'Existing Engineer', company: 'Existing Co', "
        "job id: '111'" in message
        for message in progress_messages
    )
    assert any(
        "Retrying cover letter 1/1 (attempt 1/1) - job title: 'Existing Engineer'"
        in message
        for message in progress_messages
    )


def test_regenerate_main_displays_llm_and_processed_count(
    monkeypatch,
    capsys,
):
    settings = Settings(
        llm_provider="api",
        llm_api_key="test-key",
        llm_api_model="deepseek/deepseek-v4-flash",
    )
    captured_settings: list[Settings | None] = []
    captured_modes: list[str] = []
    captured_retry_counts: list[int] = []
    captured_progress_callbacks: list[object] = []

    async def fake_run_regenerate_from_cli(
        args,
        *,
        settings=None,
        artifact_mode="resumes-only",
        progress_callback=None,
    ):
        captured_settings.append(settings)
        captured_modes.append(artifact_mode)
        captured_retry_counts.append(args.cover_letter_retries)
        captured_progress_callbacks.append(progress_callback)
        assert args.job_ids == ["111"]
        return MatchingJobsWorkflowResult(
            profile_files=[],
            search_queries=[],
            jobs_found=3,
            resumes_created=1,
            cover_letters_created=1,
            recommendations_created=1,
            tracking_spreadsheet="output/tracking/read_applications/linkedin_applications.xlsx",
            errors=["333: failed"],
            artifacts=[
                TailoredResumeArtifact(
                    job_id="111",
                    company="Acme",
                    title="Platform Engineer",
                    linkedin_url="https://www.linkedin.com/jobs/view/111",
                    resume_path="output/resumes/acme/111/resume.pdf",
                ),
                TailoredResumeArtifact(
                    job_id="222",
                    company="Beta",
                    title="Software Engineer",
                    linkedin_url="https://www.linkedin.com/jobs/view/222",
                    resume_path="output/resumes/beta/222/recommendations.pdf",
                    artifact_kind="recommendations",
                    recommendations_path="output/resumes/beta/222/recommendations.pdf",
                ),
            ],
        )

    monkeypatch.setattr(matching, "load_settings", lambda: settings)
    monkeypatch.setattr(matching, "run_regenerate_from_cli", fake_run_regenerate_from_cli)
    monkeypatch.setattr(sys, "argv", ["linkedin-career-regenerate-resumes", "111"])

    matching.regenerate_main()

    assert captured_settings == [settings]
    assert captured_modes == ["resumes-only"]
    assert captured_retry_counts == [1]
    assert captured_progress_callbacks[0] is matching._stderr_progress
    captured = capsys.readouterr()
    assert "LLM: api:deepseek/deepseek-v4-flash" in captured.err
    assert (
        "Jobs processed: 2/3 "
        "(resumes: 1, cover letters: 1, recommendations: 1, errors: 1)"
    ) in captured.err
    assert "Artifact audit: 0/0 resumes, 0/0 cover letters." in captured.err
    assert "Missing artifacts: none." in captured.err
    result = json.loads(captured.out)
    assert result["jobs_found"] == 3
    assert len(result["artifacts"]) == 2
