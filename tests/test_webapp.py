import os
import sqlite3
import threading
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from reportlab.pdfgen import canvas

from linkedin_career_mcp import webapp
from linkedin_career_mcp.webapp import create_app, import_output_artifacts


def test_connect_database_migrates_job_description_columns(tmp_path: Path):
    database_path = tmp_path / "applications.sqlite3"
    with sqlite3.connect(database_path) as connection:
        connection.execute(
            """
            CREATE TABLE applications (
                job_id TEXT PRIMARY KEY,
                company TEXT NOT NULL,
                job_title TEXT NOT NULL,
                linkedin_url TEXT NOT NULL,
                resume_filename TEXT NOT NULL,
                resume_content BLOB,
                resume_mime_type TEXT NOT NULL DEFAULT 'application/pdf',
                source_resume_path TEXT NOT NULL,
                applied_to TEXT NOT NULL DEFAULT 'No',
                date_applied TEXT,
                notes TEXT NOT NULL DEFAULT '',
                imported_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            INSERT INTO applications (
                job_id, company, job_title, linkedin_url, resume_filename,
                source_resume_path, applied_to, imported_at, updated_at
            )
            VALUES (
                '123', 'Example Co', 'Senior Engineer',
                'https://www.linkedin.com/jobs/view/123', 'resume.pdf',
                'output/resumes/resume.pdf', 'No',
                '2026-06-08T10:00:00+00:00',
                '2026-06-08T10:00:00+00:00'
            )
            """
        )
        connection.commit()

    with webapp.connect_database(database_path) as connection:
        rows = connection.execute("PRAGMA table_info(applications)").fetchall()

    columns = {row["name"] for row in rows}
    assert "job_description" in columns
    assert "prompt_job_description" in columns
    assert "application_resume_object" in columns
    assert "application_resume_updated_at" in columns
    assert "application_resume_backup_object" in columns
    assert "application_resume_backup_created_at" in columns
    assert "resume_html_filename" in columns
    assert "resume_html_content" in columns
    assert "resume_html_mime_type" in columns
    assert "source_resume_html_path" in columns
    assert "resume_html_updated_at" in columns
    assert "resume_updated_at" in columns
    assert "cover_letter_filename" in columns
    assert "cover_letter_content" in columns
    assert "source_cover_letter_path" in columns
    assert "cover_letter_updated_at" in columns
    assert "date_matched" in columns
    assert "date_posted" in columns
    assert "experience_level" in columns
    assert "ats_score" in columns
    assert "ats_parsing_score" in columns
    assert "ats_keyword_score" in columns
    assert "ats_semantic_score" in columns
    assert "ats_formatting_risk" in columns
    assert "ats_missing_terms" in columns
    with webapp.connect_database(database_path) as connection:
        row = connection.execute(
            """
            SELECT date_matched, date_posted
            FROM applications
            WHERE job_id = '123'
            """
        ).fetchone()
    assert row["date_matched"] == "2026-06-08T10:00:00+00:00"
    assert row["date_posted"] is None


def test_import_output_artifacts_stores_workbook_rows_and_artifact_blobs(
    tmp_path: Path,
    monkeypatch,
):
    output_dir = tmp_path / "output"
    resume_path = (
        output_dir
        / "resumes"
        / "Example_Co"
        / "123_senior_engineer"
        / "mp_resume_senior_engineer.pdf"
    )
    resume_path.parent.mkdir(parents=True)
    resume_path.write_bytes(b"%PDF-1.4 fake pdf")
    resume_updated_at = datetime(2026, 6, 9, 15, 30, tzinfo=UTC)
    os.utime(resume_path, (resume_updated_at.timestamp(), resume_updated_at.timestamp()))
    cover_letter_path = (
        output_dir
        / "cover_letters"
        / "Example_Co"
        / "123_senior_engineer"
        / "mp_cover_letter_senior_engineer.pdf"
    )
    cover_letter_path.parent.mkdir(parents=True)
    cover_letter_path.write_bytes(b"%PDF-1.4 fake cover")
    cover_letter_updated_at = datetime(2026, 6, 10, 16, 45, tzinfo=UTC)
    os.utime(
        cover_letter_path,
        (cover_letter_updated_at.timestamp(), cover_letter_updated_at.timestamp()),
    )

    tracking_path = output_dir / "tracking/read_applications/linkedin_applications.xlsx"
    tracking_path.parent.mkdir(parents=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Applications"
    sheet.append(
        [
            "job_id",
            "company",
            "job_title",
            "linkedin_url",
            "customized_resume",
            "cover_letter",
            "applied_to",
            "date_applied",
        ]
    )
    sheet.append(
        [
            "123",
            "Example Co",
            "Senior Engineer",
            "https://www.linkedin.com/jobs/view/123",
            str(resume_path.relative_to(tmp_path)),
            str(cover_letter_path.relative_to(tmp_path)),
            "No",
            "",
        ]
    )
    workbook.save(tracking_path)

    database_path = output_dir / "tracking/applications.sqlite3"
    result = import_output_artifacts(output_dir=output_dir, database_path=database_path)

    assert result.rows_seen == 1
    assert result.rows_imported == 1
    assert result.missing_resumes == 0
    assert result.missing_cover_letters == 0
    assert database_path.exists()
    webapp.upsert_application_artifact(
        database_path=database_path,
        job_id="123",
        company="Example Co",
        job_title="Senior Engineer",
        linkedin_url="https://www.linkedin.com/jobs/view/123",
        resume_path=resume_path,
        cover_letter_path=cover_letter_path,
        job_description="Full parsed JOD with mission boilerplate and role requirements.",
        prompt_job_description="Clean prompt JOD with role requirements.",
        date_posted="2026-06-07T12:00:00Z",
        experience_level="Mid-Senior level",
    )
    with webapp.connect_database(database_path) as connection:
        score_row = connection.execute(
            """
            SELECT ats_score, ats_parsing_score, ats_keyword_score,
                   ats_semantic_score, ats_formatting_risk, ats_missing_terms,
                   experience_level, resume_updated_at, cover_letter_updated_at
            FROM applications
            WHERE job_id = '123'
            """
        ).fetchone()
    assert score_row["ats_score"] is not None
    assert score_row["ats_formatting_risk"] in {"Low", "Medium", "High"}
    assert score_row["ats_missing_terms"] is not None
    assert score_row["experience_level"] == "Mid-Senior level"
    assert score_row["resume_updated_at"] == "2026-06-09T15:30:00+00:00"
    assert score_row["cover_letter_updated_at"] == "2026-06-10T16:45:00+00:00"

    app = create_app(database_path=database_path, output_dir=output_dir)
    client = app.test_client()

    index = client.get("/")
    assert index.status_code == 200
    html = index.data.decode()
    assert b"/descriptions/123" in index.data
    assert b"Compare descriptions" in index.data
    assert b"Cover Letter" in index.data
    assert b"/cover-letters/123" in index.data
    assert 'action="/resumes/123/copy-to-downloads"' in html
    assert 'action="/cover-letters/123/copy-to-downloads"' in html
    assert 'id="actions-form"' in html
    assert 'action="/actions/run"' in html
    assert 'id="action-sync"' in html
    assert "Sync from output" in html
    assert 'id="action-regenerate"' in html
    assert "Regenerate docs" in html
    assert "Cover letters" in html
    assert "Resumes" in html
    assert 'id="action-status"' in html
    assert 'fetch("/actions/status"' in html
    assert 'class="same-page-download"' not in html
    assert "downloadInCurrentPage" not in html
    assert "samePageDownloads.forEach" not in html
    assert 'href="/resumes/123/download"' not in html
    assert 'href="/cover-letters/123/download"' not in html
    assert b"N/A" in index.data
    assert b"Rejected" in index.data
    assert b"Accepted for interview" in index.data
    assert b"<th>Posted</th>" in index.data
    assert b"<th>Experience</th>" in index.data
    assert b"Mid-Senior level" in index.data
    assert b'id="company-sort"' in index.data
    assert b'id="matched-sort"' in index.data
    assert b'id="ats-sort"' in index.data
    assert b'id="resume-sort"' in index.data
    assert b'id="cover-letter-sort"' in index.data
    assert b'data-company-sort="Example Co"' in index.data
    assert b'data-matched-sort=' in index.data
    assert b"data-ats-sort=" in index.data
    assert b'data-resume-sort="2026-06-09T15:30:00+00:00"' in index.data
    assert b'data-cover-letter-sort="2026-06-10T16:45:00+00:00"' in index.data
    assert b"2026-06-09 15:30" in index.data
    assert b"2026-06-10 16:45" in index.data
    assert b"ATS proxy score:" in index.data
    assert b"Keyword match:" in index.data
    assert b"Formatting risk:" in index.data
    assert b"Missing/high-value terms:" in index.data
    assert b"2026-06-07" in index.data

    filtered_index = client.get(
        "/?status=Accepted+for+interview&q=Example&sort=ats&direction=desc"
    )
    filtered_html = filtered_index.data.decode()
    assert 'value="Example"' in filtered_html
    assert (
        'value="/?status=Accepted+for+interview&amp;q=Example&amp;sort=ats&amp;direction=desc"'
        in filtered_html
    )
    assert 'const initialSort = "ats";' in filtered_html
    assert 'const initialDirection = "desc";' in filtered_html
    assert "preserve-state-link" in filtered_html
    assert "return-to-state" in filtered_html

    artifact_sorted_index = client.get("/?sort=resume&direction=desc")
    artifact_sorted_html = artifact_sorted_index.data.decode()
    assert 'const initialSort = "resume";' in artifact_sorted_html
    assert 'const initialDirection = "desc";' in artifact_sorted_html

    cover_sorted_index = client.get("/?sort=cover_letter&direction=asc")
    cover_sorted_html = cover_sorted_index.data.decode()
    assert 'const initialSort = "cover_letter";' in cover_sorted_html
    assert 'const initialDirection = "asc";' in cover_sorted_html

    descriptions = client.get("/descriptions/123")
    assert descriptions.status_code == 200
    assert b"Parsed Job Description" in descriptions.data
    assert b"Prompt Job Description" in descriptions.data
    assert b"Full parsed JOD with mission boilerplate and role requirements." in descriptions.data
    assert b"Clean prompt JOD with role requirements." in descriptions.data

    db_resume = client.get("/resumes/123")
    assert db_resume.status_code == 200
    assert db_resume.data == b"%PDF-1.4 fake pdf"

    resume_download = client.get("/resumes/123/download")
    assert resume_download.status_code == 200
    assert resume_download.data == b"%PDF-1.4 fake pdf"
    assert "attachment" in resume_download.headers["Content-Disposition"]

    db_cover_letter = client.get("/cover-letters/123")
    assert db_cover_letter.status_code == 200
    assert db_cover_letter.data == b"%PDF-1.4 fake cover"

    cover_letter_download = client.get("/cover-letters/123/download")
    assert cover_letter_download.status_code == 200
    assert cover_letter_download.data == b"%PDF-1.4 fake cover"
    assert "attachment" in cover_letter_download.headers["Content-Disposition"]

    downloads_dir = tmp_path / "Downloads"
    downloads_dir.mkdir()
    downloaded_resume = downloads_dir / "mp_resume_senior_engineer.pdf"
    downloaded_cover_letter = downloads_dir / "mp_cover_letter_senior_engineer.pdf"
    unrelated_pdf = downloads_dir / "other_resume.pdf"
    monkeypatch.setenv("HOME", str(tmp_path))

    resume_copy = client.post("/resumes/123/copy-to-downloads")
    assert resume_copy.status_code == 302
    assert downloaded_resume.read_bytes() == b"%PDF-1.4 fake pdf"

    cover_letter_copy = client.post("/cover-letters/123/copy-to-downloads")
    assert cover_letter_copy.status_code == 302
    assert downloaded_cover_letter.read_bytes() == b"%PDF-1.4 fake cover"

    copy_index = client.get("/")
    assert b"Copied resume to ~/Downloads/mp_resume_senior_engineer.pdf." in copy_index.data
    assert (
        b"Copied cover letter to ~/Downloads/mp_cover_letter_senior_engineer.pdf."
        in copy_index.data
    )

    unrelated_pdf.write_bytes(b"other")

    yes_response = client.post(
        "/applications/123",
        data={"applied_to": "Yes", "date_applied": "2026-06-08", "notes": ""},
    )
    assert yes_response.status_code == 302
    assert not downloaded_resume.exists()
    assert not downloaded_cover_letter.exists()
    assert unrelated_pdf.exists()

    update_response = client.post(
        "/applications/123",
        data={"applied_to": "N/A", "date_applied": "", "notes": "Skip this one"},
    )
    assert update_response.status_code == 302
    refreshed_index = client.get("/")
    assert b'N/A: 1' in refreshed_index.data

    interview_response = client.post(
        "/applications/123",
        data={
            "applied_to": "Accepted for interview",
            "date_applied": "",
            "notes": "Recruiter screen scheduled",
            "return_to": "/?status=Accepted+for+interview&q=Example&sort=ats&direction=desc",
        },
    )
    assert interview_response.status_code == 302
    assert interview_response.headers["Location"] == (
        "/?status=Accepted+for+interview&q=Example&sort=ats&direction=desc"
    )
    interview_index = client.get("/")
    assert b"Interview: 1" in interview_index.data
    assert b'data-status="Accepted for interview"' in interview_index.data

    rejected_response = client.post(
        "/applications/123",
        data={"applied_to": "Rejected", "date_applied": "", "notes": "Closed out"},
    )
    assert rejected_response.status_code == 302
    rejected_index = client.get("/")
    assert b"Rejected: 1" in rejected_index.data
    assert b'data-status="Rejected"' in rejected_index.data

    output_resume = client.get(
        "/output/resumes/Example_Co/123_senior_engineer/mp_resume_senior_engineer.pdf"
    )
    assert output_resume.status_code == 200
    assert output_resume.data == b"%PDF-1.4 fake pdf"

    output_cover_letter = client.get(
        "/output/cover_letters/Example_Co/123_senior_engineer/mp_cover_letter_senior_engineer.pdf"
    )
    assert output_cover_letter.status_code == 200
    assert output_cover_letter.data == b"%PDF-1.4 fake cover"

    opened_urls: list[str] = []
    monkeypatch.setattr(webapp, "_open_url_in_chromium", opened_urls.append)
    linkedin_response = client.get(
        "/linkedin/123",
        query_string={"return_to": "/?status=Rejected&sort=company&direction=asc"},
    )
    assert linkedin_response.status_code == 302
    assert linkedin_response.headers["Location"] == (
        "/?status=Rejected&sort=company&direction=asc"
    )
    assert opened_urls == ["https://www.linkedin.com/jobs/view/123"]

    response = client.post("/applications/delete", data={"job_id": "123"})
    assert response.status_code == 302
    assert client.get("/resumes/123").status_code == 404
    assert client.get("/cover-letters/123").status_code == 404

    workbook = load_workbook(tracking_path)
    sheet = workbook.active
    assert sheet.max_row == 1


def test_regenerate_make_command_maps_modes_to_make_targets():
    assert webapp._regenerate_make_command(  # noqa: SLF001
        regenerate_mode="all",
        job_ids=["123", "456"],
    ) == ["make", "regenerate-all", "JOB_IDS=123 456"]
    assert webapp._regenerate_make_command(  # noqa: SLF001
        regenerate_mode="resumes",
        job_ids=["123"],
    ) == ["make", "regenerate-resumes", "JOB_IDS=123"]
    assert webapp._regenerate_make_command(  # noqa: SLF001
        regenerate_mode="cover_letters",
        job_ids=["123"],
    ) == ["make", "regenerate-cover-letters", "JOB_IDS=123"]


def test_store_application_resume_first_draft_updates_tracker_row(tmp_path: Path):
    database_path = tmp_path / "applications.sqlite3"
    webapp.upsert_application_artifact(
        database_path=database_path,
        job_id="123",
        company="Example Co",
        job_title="Senior Engineer",
        linkedin_url="https://www.linkedin.com/jobs/view/123",
        resume_path=None,
        job_description="Requires Python, AWS, APIs, and observability.",
        prompt_job_description="Requires Python, AWS, APIs, and observability.",
    )
    resume_pdf = _pdf_bytes(
        """
        Max Perkhounkov
        Professional Summary
        Core Technical Skills
        Python AWS APIs observability
        Professional Experience
        Built Python APIs on AWS with observability.
        Education
        Certifications
        """
    )

    webapp.store_application_resume_first_draft(
        database_path=database_path,
        job_id="123",
        application_resume_object="schema_version: test\n",
        resume_html="<html><body><h1>First Draft Resume</h1></body></html>",
        resume_pdf=resume_pdf,
    )

    with webapp.connect_database(database_path) as connection:
        row = connection.execute(
            """
            SELECT application_resume_object, resume_html_filename, resume_html_content,
                   resume_html_mime_type, resume_filename, resume_content,
                   source_resume_path, ats_score, ats_missing_terms
            FROM applications
            WHERE job_id = '123'
            """
        ).fetchone()
    assert row["application_resume_object"] == "schema_version: test\n"
    assert row["resume_html_filename"] == "mp_resume_senior_engineer.html"
    assert row["resume_html_content"] == "<html><body><h1>First Draft Resume</h1></body></html>"
    assert row["resume_html_mime_type"] == "text/html; charset=utf-8"
    assert row["resume_filename"] == "mp_resume_senior_engineer.pdf"
    assert row["resume_content"] == resume_pdf
    assert row["source_resume_path"] == ""
    assert row["ats_score"] is not None
    assert row["ats_missing_terms"] is not None

    app = create_app(database_path=database_path, output_dir=tmp_path / "output")
    client = app.test_client()
    html_response = client.get("/resume-html/123")
    assert html_response.status_code == 200
    assert b"First Draft Resume" in html_response.data
    assert html_response.mimetype == "text/html"
    index = client.get("/")
    assert b'href="/resume-html/123"' in index.data


def test_legacy_artifact_sync_preserves_first_draft_resume_when_aro_exists(
    tmp_path: Path,
):
    database_path = tmp_path / "applications.sqlite3"
    webapp.upsert_application_artifact(
        database_path=database_path,
        job_id="123",
        company="Example Co",
        job_title="Senior Engineer",
        linkedin_url="https://www.linkedin.com/jobs/view/123",
        resume_path=None,
        job_description="Requires Python, AWS, APIs, and observability.",
        prompt_job_description="Requires Python, AWS, APIs, and observability.",
    )
    first_draft_pdf = _pdf_bytes(
        """
        Max Perkhounkov
        Core Technical Skills
        Python AWS APIs observability
        Professional Experience
        Built Python APIs on AWS with observability.
        """
    )
    first_draft_path = tmp_path / "first_draft.pdf"
    first_draft_path.write_bytes(first_draft_pdf)
    webapp.store_application_resume_first_draft(
        database_path=database_path,
        job_id="123",
        application_resume_object="schema_version: test\n",
        resume_html="<html><body><h1>First Draft Resume</h1></body></html>",
        resume_pdf=first_draft_pdf,
        resume_pdf_path=first_draft_path,
    )
    with webapp.connect_database(database_path) as connection:
        first_draft_row = connection.execute(
            """
            SELECT resume_filename, resume_content, source_resume_path,
                   resume_updated_at, ats_score, ats_updated_at
            FROM applications
            WHERE job_id = '123'
            """
        ).fetchone()

    legacy_resume_path = tmp_path / "output/resumes/Example_Co/123/resume.pdf"
    legacy_resume_path.parent.mkdir(parents=True)
    legacy_resume_path.write_bytes(_pdf_bytes("Legacy Java resume"))
    legacy_updated_at = datetime(2026, 6, 9, 15, 30, tzinfo=UTC)
    os.utime(
        legacy_resume_path,
        (legacy_updated_at.timestamp(), legacy_updated_at.timestamp()),
    )
    webapp.upsert_application_artifact(
        database_path=database_path,
        job_id="123",
        company="Example Co",
        job_title="Senior Engineer",
        linkedin_url="https://www.linkedin.com/jobs/view/123",
        resume_path=legacy_resume_path,
        job_description="Requires Java and legacy systems.",
        prompt_job_description="Requires Java and legacy systems.",
    )

    with webapp.connect_database(database_path) as connection:
        row = connection.execute(
            """
            SELECT resume_filename, resume_content, source_resume_path,
                   resume_updated_at, ats_score, ats_updated_at
            FROM applications
            WHERE job_id = '123'
            """
        ).fetchone()
    assert row["resume_filename"] == first_draft_row["resume_filename"]
    assert row["resume_content"] == first_draft_row["resume_content"]
    assert row["source_resume_path"] == str(first_draft_path)
    assert row["resume_updated_at"] == first_draft_row["resume_updated_at"]
    assert row["ats_score"] == first_draft_row["ats_score"]
    assert row["ats_updated_at"] == first_draft_row["ats_updated_at"]


def test_resume_editor_saves_rerenders_rescores_and_reverts(tmp_path: Path, monkeypatch):
    database_path = tmp_path / "applications.sqlite3"
    webapp.upsert_application_artifact(
        database_path=database_path,
        job_id="123",
        company="Example Co",
        job_title="Senior Engineer",
        linkedin_url="https://www.linkedin.com/jobs/view/123",
        resume_path=None,
        job_description="Requires Python, AWS, APIs, and authentication.",
        prompt_job_description="Requires Python, AWS, APIs, and authentication.",
    )
    original_aro = _sample_application_resume_yaml(
        paragraph="Original summary for platform APIs.",
        bullet="Original platform API bullet.",
    )
    webapp.store_application_resume_first_draft(
        database_path=database_path,
        job_id="123",
        application_resume_object=original_aro,
        resume_html="<html><body><h1>Original</h1></body></html>",
        resume_pdf=_pdf_bytes("Original Python AWS APIs"),
    )

    monkeypatch.setattr(
        webapp,
        "render_resume_pdf_from_html",
        lambda html: _pdf_bytes(f"Rendered {html}"),
    )

    app = create_app(database_path=database_path, output_dir=tmp_path / "output")
    client = app.test_client()
    index = client.get("/")
    assert b"/resumes/123/edit" in index.data

    edit_response = client.get("/resumes/123/edit?return_to=%2F%3Fsort%3Dresume")
    assert edit_response.status_code == 200
    form_data = _edit_form_data(edit_response.data.decode())
    form_data["summary_paragraph"] = "Edited summary with authentication APIs."
    form_data["job_0_bullet_0_text"] = "Edited authentication API bullet."

    save_response = client.post("/resumes/123/edit", data=form_data)
    assert save_response.status_code == 302
    with webapp.connect_database(database_path) as connection:
        edited_row = connection.execute(
            """
            SELECT application_resume_object, application_resume_backup_object,
                   resume_html_content, resume_content, source_resume_path,
                   ats_score
            FROM applications
            WHERE job_id = '123'
            """
        ).fetchone()
    assert "Edited summary with authentication APIs." in edited_row[
        "application_resume_object"
    ]
    assert "Edited authentication API bullet." in edited_row["application_resume_object"]
    assert "Original platform API bullet." in edited_row[
        "application_resume_backup_object"
    ]
    assert "Edited authentication API bullet." in edited_row["resume_html_content"]
    assert edited_row["resume_content"] is not None
    assert edited_row["source_resume_path"] == ""
    assert edited_row["ats_score"] is not None

    revert_response = client.post(
        "/resumes/123/edit/revert",
        data={"return_to": "/?sort=resume"},
    )
    assert revert_response.status_code == 302
    with webapp.connect_database(database_path) as connection:
        reverted_row = connection.execute(
            """
            SELECT application_resume_object, application_resume_backup_object,
                   resume_html_content, ats_score
            FROM applications
            WHERE job_id = '123'
            """
        ).fetchone()
    assert "Original platform API bullet." in reverted_row["application_resume_object"]
    assert "Edited authentication API bullet." in reverted_row[
        "application_resume_backup_object"
    ]
    assert "Original platform API bullet." in reverted_row["resume_html_content"]
    assert reverted_row["ats_score"] is not None


def test_actions_run_starts_background_regeneration(tmp_path: Path):
    with webapp._ACTION_RUN_LOCK:  # noqa: SLF001
        webapp._ACTION_RUNS.clear()  # noqa: SLF001

    calls = []
    completed = threading.Event()

    def runner(**kwargs):
        calls.append(kwargs)
        webapp._append_background_action_message(  # noqa: SLF001
            kwargs["run_id"],
            "processing job 123",
        )
        webapp._finish_background_action_run(  # noqa: SLF001
            kwargs["run_id"],
            status="completed",
            return_code=0,
        )
        completed.set()

    output_dir = tmp_path / "output"
    database_path = output_dir / "tracking/applications.sqlite3"
    app = create_app(
        database_path=database_path,
        output_dir=output_dir,
        background_action_runner=runner,
    )
    client = app.test_client()

    response = client.post(
        "/actions/run",
        data={
            "action_sync": "1",
            "action_regenerate": "1",
            "regenerate_mode": "all",
            "job_id": ["123", "123", "456"],
            "return_to": "/?q=Example",
        },
    )

    assert response.status_code == 302
    assert response.headers["Location"] == "/?q=Example"
    assert completed.wait(timeout=2)
    assert calls
    assert calls[0]["sync_requested"] is True
    assert calls[0]["regenerate_mode"] == "all"
    assert calls[0]["job_ids"] == ["123", "456"]

    status = client.get("/actions/status").get_json()
    assert status is not None
    assert status["runs"][0]["status"] == "completed"
    assert status["runs"][0]["return_code"] == 0
    assert status["runs"][0]["title"] == "sync output + regenerate all docs for 2 job(s)"
    assert any("processing job 123" in message for message in status["runs"][0]["messages"])


def _edit_form_data(html: str) -> dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    form = soup.find("form", {"action": "/resumes/123/edit"})
    assert form is not None
    data: dict[str, str] = {}
    for field in form.find_all(["input", "textarea"]):
        name = field.get("name")
        if not name:
            continue
        if field.name == "textarea":
            data[name] = field.text
            continue
        field_type = str(field.get("type") or "text").lower()
        if field_type == "checkbox":
            if field.has_attr("checked"):
                data[name] = str(field.get("value") or "on")
            continue
        data[name] = str(field.get("value") or "")
    return data


def _sample_application_resume_yaml(*, paragraph: str, bullet: str) -> str:
    return f"""schema_version: test
header_top:
  line_1_name_header_text: Max Perkhounkov
  line_2_applicant_info_text: max@example.com
  contact_items:
  - max@example.com
  links: []
professional_summary:
  render: true
  header_text: Professional Summary
  paragraph: {paragraph}
  summary_note: ''
core_technical_skills:
  render: true
  header_text: Core Technical Skills
  bullet_points:
  - order: 1
    category: Platform Engineering
    items:
      primary:
      - Python
      - AWS
      additional:
      - authentication
      - Java
    jod_matched_items:
    - authentication
professional_experience:
  render: true
  header_text: Professional Experience
  jobs:
  - order: 1
    render: true
    line_1:
      company_name_text: Example Co
      position_name_text: Senior Engineer
      position_dates_text: 2020 - Present
    line_2:
      position_intro_text: Platform engineering role.
    bullet_points:
    - order: 1
      render: true
      text: {bullet}
education:
  render: true
  header_text: Education
  entries:
  - order: 1
    render: true
    line_1:
      institution_name_text: University
    line_2:
      degree_name_text: BS Physics
      degree_dates_text: 2009 - 2013
    bullet_points:
    - order: 1
      render: true
      text: Applied math coursework.
certifications:
  render: true
  header_text: Certifications
  bullet_points:
  - order: 1
    render: true
    text: OCI Engineer
portfolio:
  render: true
  header_text: Portfolio
  projects:
  - order: 1
    render: true
    title_text: LinkedIn Career MCP
    url: https://github.com/mperkhou/linkedin-career-mcp
    description_text: Resume automation workflow.
"""


def _pdf_bytes(text: str) -> bytes:
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer)
    text_object = pdf.beginText(40, 760)
    for line in text.strip().splitlines():
        text_object.textLine(line.strip())
    pdf.drawText(text_object)
    pdf.save()
    return buffer.getvalue()
